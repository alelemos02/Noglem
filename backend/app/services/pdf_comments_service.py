import io
import re
import tempfile
from collections import Counter
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional

import fitz  # PyMuPDF
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

COMMENT_TYPES = {"Text", "FreeText", "Highlight", "Underline", "StrikeOut", "Squiggly"}
MARKUP_TYPES = {"Highlight", "Underline", "StrikeOut", "Squiggly"}
_PDF_FLAG_LOCKED = 1 << 7  # = 128 — campos de template CAD (AutoCAD/Revit)


# ── Data models ───────────────────────────────────────────────────────────────

@dataclass
class Annotation:
    document_number: str
    page: int
    annotation_type: str
    author: str
    date: str
    comment: str
    marked_text: str
    subject: str
    ai_analysis: str = ""


@dataclass
class ExtractionResult:
    filename: str
    annotations: list = field(default_factory=list)
    error: Optional[str] = None
    page_count: int = 0


# ── Helpers ───────────────────────────────────────────────────────────────────

def _parse_pdf_date(raw: str) -> str:
    if not raw:
        return ""
    m = re.match(r"D:(\d{4})(\d{2})(\d{2})(\d{2})(\d{2})(\d{2})", raw)
    if not m:
        return raw.strip()
    y, mo, d, h, mi, _ = m.groups()
    return f"{d}/{mo}/{y} {h}:{mi}"


def _extract_document_number(filename: str) -> str:
    return Path(filename).stem


def _is_template_field(annot) -> bool:
    return bool(annot.flags & _PDF_FLAG_LOCKED)


# ── Extractor ─────────────────────────────────────────────────────────────────

def extract_from_pdf(path: Path) -> ExtractionResult:
    filename = path.name
    document_number = _extract_document_number(filename)

    try:
        doc = fitz.open(str(path))
    except Exception as exc:
        return ExtractionResult(filename=filename, error=f"Falha ao abrir PDF: {exc}")

    if doc.is_encrypted:
        if not doc.authenticate(""):
            doc.close()
            return ExtractionResult(filename=filename, error="PDF protegido por senha")

    page_count = len(doc)
    annotations: list[Annotation] = []

    try:
        for page_index in range(page_count):
            page = doc[page_index]
            page_number = page_index + 1

            for annot in page.annots():
                annot_type = annot.type[1]
                if annot_type not in COMMENT_TYPES:
                    continue

                info = annot.info
                comment = info.get("content", "").strip()
                author = info.get("title", "").strip()
                subject = info.get("subject", "").strip()
                raw_date = info.get("modDate", "") or info.get("creationDate", "")
                date = _parse_pdf_date(raw_date)

                if annot_type == "FreeText" and _is_template_field(annot):
                    continue

                marked_text = ""
                if annot_type in MARKUP_TYPES:
                    try:
                        marked_text = page.get_text("text", clip=annot.rect).strip()
                    except Exception:
                        marked_text = ""

                if not comment and not marked_text:
                    continue

                annotations.append(Annotation(
                    document_number=document_number,
                    page=page_number,
                    annotation_type=annot_type,
                    author=author,
                    date=date,
                    comment=comment,
                    marked_text=marked_text,
                    subject=subject,
                ))
    finally:
        doc.close()

    return ExtractionResult(
        filename=filename,
        annotations=[asdict(a) for a in annotations],
        page_count=page_count,
    )


# ── Excel export ─────────────────────────────────────────────────────────────

HEADER_BG = "1E3A5F"
HEADER_FG = "FFFFFF"
ROW_ODD_BG = "FFFFFF"
ROW_EVEN_BG = "F0F5FA"
AI_BG = "EBF3FE"
AI_BG_EVEN = "DDE9F8"
BORDER_COLOR = "D0D8E4"

COLUMNS = [
    ("Documento",  42, "document_number"),
    ("Página",      8, "page"),
    ("Autor",      22, "author"),
    ("Data/Hora",  17, "date"),
    ("Comentário", 65, "comment"),
    ("Análise IA", 78, "ai_analysis"),
]


def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _bottom_border() -> Border:
    side = Side(style="thin", color=BORDER_COLOR)
    return Border(bottom=side)


def _header_border() -> Border:
    side = Side(style="medium", color="FFFFFF")
    return Border(bottom=side)


def _build_comments_sheet(ws, annotations: list[dict]) -> None:
    ws.title = "Comentários"
    ws.sheet_view.showGridLines = False

    header_font = Font(bold=True, color=HEADER_FG, size=11, name="Calibri")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    ws.row_dimensions[1].height = 36

    for col_idx, (label, width, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = _fill(HEADER_BG)
        cell.alignment = header_align
        cell.border = _header_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    data_font = Font(size=10, name="Calibri")
    ai_font = Font(size=10, name="Calibri", italic=True, color="1E3A5F")
    page_align = Alignment(horizontal="center", vertical="top")
    left_align = Alignment(horizontal="left", vertical="top")
    wrap_align = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for row_idx, annot in enumerate(annotations, start=2):
        is_even = row_idx % 2 == 0
        row_bg = ROW_EVEN_BG if is_even else ROW_ODD_BG
        ai_bg = AI_BG_EVEN if is_even else AI_BG
        ws.row_dimensions[row_idx].height = 15

        for col_idx, (_, _, field_key) in enumerate(COLUMNS, start=1):
            value = annot.get(field_key, "") or ""
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = _bottom_border()

            if field_key == "ai_analysis":
                cell.fill = _fill(ai_bg)
                cell.font = ai_font
                cell.alignment = wrap_align
            elif field_key == "comment":
                cell.fill = _fill(row_bg)
                cell.font = data_font
                cell.alignment = wrap_align
            elif field_key == "page":
                cell.fill = _fill(row_bg)
                cell.font = data_font
                cell.alignment = page_align
            else:
                cell.fill = _fill(row_bg)
                cell.font = data_font
                cell.alignment = left_align

    last_col = get_column_letter(len(COLUMNS))
    ws.auto_filter.ref = f"A1:{last_col}{max(len(annotations) + 1, 1)}"
    ws.freeze_panes = "A2"


def _build_summary_sheet(ws, annotations: list[dict], errors: list[dict]) -> None:
    ws.title = "Resumo"
    ws.sheet_view.showGridLines = False

    header_font = Font(bold=True, color=HEADER_FG, size=11, name="Calibri")
    label_font = Font(size=10, name="Calibri")
    value_font = Font(size=10, name="Calibri", bold=True)
    total = len(annotations)
    row = 1

    def _write_section_header(label: str):
        nonlocal row
        for c in range(1, 3):
            cell = ws.cell(row=row, column=c, value=label if c == 1 else "")
            cell.font = header_font
            cell.fill = _fill(HEADER_BG)
            cell.border = _header_border()
            cell.alignment = Alignment(vertical="center", horizontal="left" if c == 1 else "center")
        ws.row_dimensions[row].height = 28
        row += 1

    def _write_row(label: str, value, even: bool):
        nonlocal row
        bg = ROW_EVEN_BG if even else ROW_ODD_BG
        c1 = ws.cell(row=row, column=1, value=label)
        c1.font = label_font
        c1.fill = _fill(bg)
        c1.border = _bottom_border()
        c1.alignment = Alignment(vertical="center", horizontal="left")

        c2 = ws.cell(row=row, column=2, value=value)
        c2.font = value_font
        c2.fill = _fill(bg)
        c2.border = _bottom_border()
        c2.alignment = Alignment(vertical="center", horizontal="center")
        ws.row_dimensions[row].height = 20
        row += 1

    _write_section_header("Informações Gerais")
    _write_row("Total de comentários", total, False)
    _write_row("Total de documentos", len({a.get("document_number", "") for a in annotations}), True)
    _write_row("Total de erros", len(errors), False)
    _write_row("Gerado em", datetime.now().strftime("%d/%m/%Y %H:%M"), True)
    row += 1

    _write_section_header("Comentários por Documento")
    by_doc = Counter(a.get("document_number", "") for a in annotations)
    for i, (doc, count) in enumerate(sorted(by_doc.items())):
        _write_row(doc, count, i % 2 == 0)
    row += 1

    _write_section_header("Comentários por Autor")
    by_author = Counter(a.get("author", "") or "(sem autor)" for a in annotations)
    for i, (author, count) in enumerate(sorted(by_author.items())):
        _write_row(author, count, i % 2 == 0)

    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 18
    ws.freeze_panes = "A2"


def _build_errors_sheet(ws, errors: list[dict]) -> None:
    ws.title = "Erros"
    ws.sheet_view.showGridLines = False

    header_font = Font(bold=True, color=HEADER_FG, size=11, name="Calibri")
    for col, label in enumerate(["Arquivo", "Motivo do Erro"], start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = header_font
        cell.fill = _fill(HEADER_BG)
        cell.border = _header_border()
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, err in enumerate(errors, start=2):
        for c, val in enumerate([err.get("filename", ""), err.get("error", "")], start=1):
            cell = ws.cell(row=row_idx, column=c, value=val)
            cell.fill = _fill("FCE4D6")
            cell.border = _bottom_border()
            cell.font = Font(size=10, name="Calibri")
            cell.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 65


def generate_excel(results: list[dict]) -> io.BytesIO:
    wb = Workbook()
    wb.remove(wb.active)

    annotations: list[dict] = []
    errors: list[dict] = []

    for result in results:
        if result.get("error"):
            errors.append({"filename": result["filename"], "error": result["error"]})
        else:
            annotations.extend(result.get("annotations", []))

    ws_comments = wb.create_sheet()
    _build_comments_sheet(ws_comments, annotations)

    ws_summary = wb.create_sheet()
    _build_summary_sheet(ws_summary, annotations, errors)

    if errors:
        ws_errors = wb.create_sheet()
        _build_errors_sheet(ws_errors, errors)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

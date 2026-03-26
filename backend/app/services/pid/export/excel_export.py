"""Export extraction results to Excel (.xlsx) with professional formatting."""

import logging
from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.services.pid.models.instrument import ExtractionResult

logger = logging.getLogger(__name__)

# Style definitions
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

DATA_FONT = Font(name="Calibri", size=10)
DATA_ALIGNMENT = Alignment(vertical="center")

WARNING_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
ERROR_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def export_to_excel(result: ExtractionResult, output_path: str) -> str:
    """Export extraction results to a formatted Excel file.

    Creates four sheets:
    1. Instrument Index - Main tag list
    2. Loops - Loop summary
    3. Validation - Warnings and errors
    4. Drawing Info - Title block metadata
    """
    wb = Workbook()

    _create_instrument_sheet(wb, result)
    _create_loops_sheet(wb, result)
    _create_validation_sheet(wb, result)
    _create_metadata_sheet(wb, result)

    path = Path(output_path)
    wb.save(str(path))
    logger.info(f"Excel exported to {path.absolute()}")
    return str(path.absolute())


def _create_instrument_sheet(wb: Workbook, result: ExtractionResult) -> None:
    """Create the main Instrument Index sheet."""
    ws = wb.active
    ws.title = "Instrument Index"

    headers = [
        "Tag Number",
        "ISA Type",
        "Description",
        "Symbol",
        "Classification",
        "Physical?",
        "Fornecido Pacote?",
        "Area",
        "Tag Number (Num)",
        "Qualifier",
        "Equipment",
        "Loop ID",
        "Line Number",
        "Service",
        "Parent Tag",
        "Children",
        "Sheet",
        "Confidence",
        "Notes",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    sorted_instruments = sorted(
        result.instruments,
        key=lambda i: (i.area or "", i.isa_type, i.tag_number or "", i.qualifier),
    )

    for row_idx, inst in enumerate(sorted_instruments, 2):
        data = [
            inst.tag,
            inst.isa_type,
            inst.isa_description,
            inst.symbol,
            inst.classification,
            "Yes" if inst.is_physical else "No",
            "Yes" if getattr(inst, "furnished_by_package", False) else "No",
            inst.area,
            inst.tag_number,
            inst.qualifier,
            inst.equipment_ref,
            inst.loop_id,
            inst.line_number,
            inst.service,
            inst.parent_tag,
            ", ".join(inst.children_tags) if inst.children_tags else "",
            inst.sheet_name or str(inst.page_index + 1),
            f"{inst.confidence:.0%}",
            "; ".join(inst.notes) if inst.notes else "",
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGNMENT
            cell.border = THIN_BORDER

            # Highlight low confidence
            if col == 18 and inst.confidence < 0.5:
                cell.fill = WARNING_FILL

            # Color Physical vs DCS
            if col == 6:
                if inst.is_physical:
                    cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")

    col_widths = [22, 10, 35, 12, 22, 10, 8, 12, 8, 15, 12, 30, 20, 22, 30, 8, 10, 40]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(sorted_instruments) + 1}"
    ws.freeze_panes = "A2"


def _create_loops_sheet(wb: Workbook, result: ExtractionResult) -> None:
    """Create the Loops summary sheet."""
    ws = wb.create_sheet("Loops")

    headers = ["Loop ID", "Instruments", "Complete?", "Missing"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    sorted_loops = sorted(result.loops, key=lambda l: l.loop_id)

    for row_idx, loop in enumerate(sorted_loops, 2):
        data = [
            loop.loop_id,
            ", ".join(loop.instruments),
            "Yes" if loop.is_complete else "No",
            ", ".join(loop.missing) if loop.missing else "",
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGNMENT
            cell.border = THIN_BORDER

            if col == 3 and not loop.is_complete:
                cell.fill = WARNING_FILL

    col_widths = [15, 60, 12, 40]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.auto_filter.ref = f"A1:D{len(sorted_loops) + 1}"
    ws.freeze_panes = "A2"


def _create_validation_sheet(wb: Workbook, result: ExtractionResult) -> None:
    """Create the Validation report sheet."""
    ws = wb.create_sheet("Validation")

    headers = ["Type", "Message"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    row = 2

    ws.cell(row=row, column=1, value="SUMMARY").font = Font(bold=True)
    ws.cell(
        row=row, column=2,
        value=f"{len(result.instruments)} instruments, "
              f"{len(result.loops)} loops, "
              f"{len(result.warnings)} warnings, "
              f"{len(result.errors)} errors"
    ).font = Font(bold=True)
    row += 1

    for error in result.errors:
        cell_type = ws.cell(row=row, column=1, value="ERROR")
        cell_type.fill = ERROR_FILL
        cell_type.border = THIN_BORDER
        cell_msg = ws.cell(row=row, column=2, value=error)
        cell_msg.fill = ERROR_FILL
        cell_msg.border = THIN_BORDER
        row += 1

    for warning in result.warnings:
        cell_type = ws.cell(row=row, column=1, value="WARNING")
        cell_type.fill = WARNING_FILL
        cell_type.border = THIN_BORDER
        cell_msg = ws.cell(row=row, column=2, value=warning)
        cell_msg.fill = WARNING_FILL
        cell_msg.border = THIN_BORDER
        row += 1

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 100
    ws.freeze_panes = "A2"


def _create_metadata_sheet(wb: Workbook, result: ExtractionResult) -> None:
    """Create the Drawing Info sheet with title block metadata."""
    if not result.metadata:
        return

    ws = wb.create_sheet("Drawing Info")

    headers = [
        "Document Number", "Revision", "Title", "Area",
        "Sheet", "Total Sheets", "Date", "Scale",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    for row_idx, meta in enumerate(result.metadata, 2):
        data = [
            meta.document_number,
            meta.revision,
            meta.title,
            meta.area,
            meta.sheet_number,
            meta.total_sheets,
            meta.date,
            meta.scale,
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER

    col_widths = [20, 10, 50, 10, 8, 10, 15, 10]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A2"

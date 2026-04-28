"""Geração do arquivo Excel formatado com 15 colunas de quantitativos."""

from __future__ import annotations

import io
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from .models import ElementoFundacao, ResultadoQuantitativo

COLUNAS: list[tuple[str, int]] = [
    ("ITENS",                    30),
    ("QTDE",                     8),
    ("RAIO\n(m)",                11),
    ("LARGURA\n(m)",             12),
    ("COMPRIMENTO\n(m)",         14),
    ("ALTURA\n(m)",              11),
    ("ALTURA DE\nESCAVAÇÃO (m)", 16),
    ("CONCRETO\nESTR (m³)",      14),
    ("FORMAS\nIN SITU (m²)",     14),
    ("GROUT\n(m³)",              11),
    ("C. MAGRO\n(m³)",           13),
    ("ESCAV\nh<1,2 (m³)",        14),
    ("REAT\nh<1,2 (m³)",         14),
    ("BOTA FORA\n(m³)",          13),
    ("ESTACAS\n(m)",             11),
]

COLUNA_ATRIBUTO: list[Optional[str]] = [
    None,
    "quantidade",
    "raio",
    "largura",
    "comprimento",
    "altura",
    "altura_escavacao",
    "concreto_estr",
    "formas_in_situ",
    "grout",
    "c_magro",
    "escav_h_menor_1_2",
    "reat_h_menor_1_2",
    "bota_fora",
    "estacas",
]

COLUNA_TOTAL_KEY: list[Optional[str]] = [
    None, None, None, None, None, None, None,
    "concreto_estr",
    "formas_in_situ",
    "grout",
    "c_magro",
    "escav_h_menor_1_2",
    "reat_h_menor_1_2",
    "bota_fora",
    "estacas",
]

COLUNAS_DIMENSAO = {2, 3, 4, 5, 6, 7}
COLUNAS_CALCULADAS = {8, 9, 10, 11, 12, 13, 14, 15}

COR_NAVY       = "1F4E79"
COR_AZUL_MED   = "2E75B6"
COR_AZUL_CLARO = "DEEAF1"
COR_VERDE_BG   = "E2EFDA"
COR_VERDE_FG   = "375623"
COR_BRANCO     = "FFFFFF"
COR_PRETO      = "000000"

FILL_TITULO   = PatternFill("solid", fgColor=COR_NAVY)
FILL_HEADER   = PatternFill("solid", fgColor=COR_AZUL_MED)
FILL_BRANCO   = PatternFill("solid", fgColor=COR_BRANCO)
FILL_ZEBRA    = PatternFill("solid", fgColor=COR_AZUL_CLARO)
FILL_TOTAL_1  = PatternFill("solid", fgColor=COR_VERDE_BG)
FILL_TOTAL    = PatternFill("solid", fgColor=COR_NAVY)

FONTE_TITULO  = Font(name="Calibri", size=12, bold=True,  color=COR_BRANCO)
FONTE_HEADER  = Font(name="Calibri", size=9,  bold=True,  color=COR_BRANCO)
FONTE_ITEM    = Font(name="Calibri", size=9,  bold=False, color=COR_PRETO)
FONTE_VALOR   = Font(name="Calibri", size=9,  bold=False, color=COR_NAVY)
FONTE_TOTAL_1 = Font(name="Calibri", size=9,  bold=True,  color=COR_VERDE_FG)
FONTE_TOTAL   = Font(name="Calibri", size=9,  bold=True,  color=COR_BRANCO)

_FINA    = Side(style="thin",   color="BFBFBF")
_GROSSA  = Side(style="medium", color=COR_NAVY)

BORDA_INTERNA = Border(left=_FINA, right=_FINA, top=_FINA, bottom=_FINA)
BORDA_CABEC   = Border(left=_FINA, right=_FINA, top=_GROSSA, bottom=_GROSSA)
BORDA_TOTAL_1 = Border(left=_FINA, right=_FINA, top=_GROSSA, bottom=_FINA)
BORDA_TOTAL   = Border(left=_FINA, right=_FINA, top=_FINA,   bottom=_GROSSA)

ALIN_ESQUERDA = Alignment(horizontal="left",   vertical="center", wrap_text=True)
ALIN_CENTRO   = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIN_DIREITA  = Alignment(horizontal="right",  vertical="center", wrap_text=False)

FMT_4 = '#,##0.0000'
FMT_INT = '#,##0'


def gerar_excel(resultados: list[ResultadoQuantitativo], destino: Path) -> None:
    wb = _construir_workbook(resultados)
    destino.parent.mkdir(parents=True, exist_ok=True)
    wb.save(destino)


def gerar_excel_bytes(resultados: list[ResultadoQuantitativo]) -> bytes:
    wb = _construir_workbook(resultados)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _construir_workbook(resultados: list[ResultadoQuantitativo]) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    for resultado in resultados:
        _adicionar_aba(wb, resultado)
    return wb


def _adicionar_aba(wb: Workbook, resultado: ResultadoQuantitativo) -> None:
    geo = resultado.geometria
    ws = wb.create_sheet(title=_nome_aba(geo.documento))

    n = len(COLUNAS)
    ws.merge_cells(f"A1:{get_column_letter(n)}1")
    tanques_str = " / ".join(geo.tanques) if geo.tanques else geo.documento
    ws["A1"].value = f"QUANTITATIVOS DE FUNDAÇÃO  ·  {tanques_str}"
    _aplicar(ws["A1"], fonte=FONTE_TITULO, fill=FILL_TITULO,
             alinhamento=ALIN_ESQUERDA,
             borda=Border(left=_GROSSA, right=_GROSSA, top=_GROSSA, bottom=_FINA))
    ws.row_dimensions[1].height = 30

    ws.merge_cells(f"A2:{get_column_letter(n)}2")
    ws["A2"].value = f"Documento: {geo.documento}    |    Total de tanques: {geo.total_tanques}"
    _aplicar(ws["A2"],
             fonte=Font(name="Calibri", size=8, italic=True, color="9DC3E6"),
             fill=FILL_TITULO,
             alinhamento=ALIN_ESQUERDA,
             borda=Border(left=_GROSSA, right=_GROSSA, top=_FINA, bottom=_GROSSA))
    ws.row_dimensions[2].height = 16

    for col, (nome, _) in enumerate(COLUNAS, start=1):
        cel = ws.cell(row=3, column=col, value=nome)
        _aplicar(cel, fonte=FONTE_HEADER, fill=FILL_HEADER,
                 alinhamento=ALIN_CENTRO, borda=BORDA_CABEC)
    ws.row_dimensions[3].height = 38

    linha = 4
    for i, item in enumerate(geo.itens):
        fill = FILL_ZEBRA if i % 2 == 1 else FILL_BRANCO
        _escrever_linha_item(ws, linha, item, fill)
        ws.row_dimensions[linha].height = 16
        linha += 1

    _escrever_linha_total(ws, linha, "TOTAL PARA 1 TANQUE",
                          resultado.total_1_tanque,
                          FONTE_TOTAL_1, FILL_TOTAL_1, BORDA_TOTAL_1)
    ws.row_dimensions[linha].height = 18
    linha += 1

    _escrever_linha_total(ws, linha, f"TOTAL  ({geo.total_tanques} TANQUES)",
                          resultado.total_geral,
                          FONTE_TOTAL, FILL_TOTAL, BORDA_TOTAL)
    ws.row_dimensions[linha].height = 20

    _ajustar_larguras(ws)
    ws.freeze_panes = "A4"


def _escrever_linha_item(ws, linha: int, item: ElementoFundacao, fill: PatternFill) -> None:
    for col, atributo in enumerate(COLUNA_ATRIBUTO, start=1):
        valor = item.item if atributo is None else getattr(item, atributo, None)
        cel = ws.cell(row=linha, column=col, value=valor)

        if col == 1:
            _aplicar(cel, fonte=FONTE_ITEM, fill=fill,
                     alinhamento=ALIN_ESQUERDA, borda=BORDA_INTERNA)
        elif col in COLUNAS_DIMENSAO:
            _aplicar(cel, fonte=FONTE_VALOR, fill=fill,
                     alinhamento=ALIN_CENTRO, borda=BORDA_INTERNA)
            if isinstance(valor, float):
                cel.number_format = FMT_4
            elif isinstance(valor, int):
                cel.number_format = FMT_INT
        else:
            _aplicar(cel, fonte=FONTE_VALOR, fill=fill,
                     alinhamento=ALIN_DIREITA, borda=BORDA_INTERNA)
            if isinstance(valor, float) and valor != 0:
                cel.number_format = FMT_4
            elif isinstance(valor, float) and valor == 0:
                cel.value = None


def _escrever_linha_total(
    ws, linha: int, rotulo: str, totais: dict[str, float],
    fonte: Font, fill: PatternFill, borda: Border
) -> None:
    for col, chave in enumerate(COLUNA_TOTAL_KEY, start=1):
        if col == 1:
            valor: object = rotulo
            alin = ALIN_ESQUERDA
        elif chave is None:
            valor = None
            alin = ALIN_CENTRO
        else:
            v = totais.get(chave)
            valor = v if v else None
            alin = ALIN_DIREITA

        cel = ws.cell(row=linha, column=col, value=valor)
        _aplicar(cel, fonte=fonte, fill=fill, alinhamento=alin, borda=borda)
        if isinstance(valor, float) and valor:
            cel.number_format = FMT_4


def _aplicar(cel, *, fonte, fill, alinhamento, borda) -> None:
    cel.font = fonte
    cel.fill = fill
    cel.alignment = alinhamento
    cel.border = borda


def _ajustar_larguras(ws) -> None:
    for col, (_, min_w) in enumerate(COLUNAS, start=1):
        letra = get_column_letter(col)
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cel in row:
                if cel.value:
                    texto = str(cel.value).split("\n")[0]
                    max_len = max(max_len, len(texto))
        ws.column_dimensions[letra].width = max(min_w, max_len + 2)


def _nome_aba(documento: str) -> str:
    nome = documento.replace("/", "-").replace("\\", "-").replace("*", "")
    return nome[:31]

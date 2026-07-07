import io
import uuid
import zipfile
from dataclasses import dataclass, field

from openpyxl import Workbook, load_workbook

from app.services.exporter import (
    _CARTA_COL_ITEM_ID,
    _CARTA_COL_RESPOSTA,
    _CARTA_HEADERS,
    _build_observacao,
    export_carta_pendencias,
    export_docx,
    export_pdf,
    export_xlsx,
)


@dataclass
class DummyParecer:
    numero_parecer: str = "PYTEST-EXP-1"
    projeto: str = "Projeto Teste"
    fornecedor: str = "Fornecedor X"
    revisao: str = "A"
    idioma_relatorio: str = "pt"
    parecer_geral: str | None = "APROVADO"
    status_processamento: str = "concluido"
    comentario_geral: str | None = "Comentario geral"
    conclusao: str | None = "Conclusao final"
    total_itens: int = 1
    total_aprovados: int = 1
    total_aprovados_comentarios: int = 0
    total_rejeitados: int = 0
    total_info_ausente: int = 0
    total_itens_adicionais: int = 0


@dataclass
class DummyItem:
    numero: int = 1
    categoria: str | None = "Tecnica"
    descricao_requisito: str = "Descricao requisito"
    referencia_engenharia: str | None = "Ref Eng"
    referencia_fornecedor: str | None = "Ref Forn"
    valor_requerido: str | None = "10bar"
    valor_fornecedor: str | None = "10bar"
    status: str = "A"
    justificativa_tecnica: str = "Justificativa"
    acao_requerida: str | None = "Nenhuma"
    prioridade: str | None = "BAIXA"
    norma_referencia: str | None = "NBR 1"


@dataclass
class DummyRecomendacao:
    ordem: int = 1
    texto: str = "Recomendacao teste"


def test_export_pdf_returns_pdf_bytes():
    payload = export_pdf(DummyParecer(), [DummyItem()], [DummyRecomendacao()])
    assert payload.startswith(b"%PDF")
    assert len(payload) > 1000


def test_export_xlsx_has_expected_sheets():
    payload = export_xlsx(DummyParecer(), [DummyItem()], [DummyRecomendacao()])
    wb = load_workbook(io.BytesIO(payload), data_only=True)
    assert wb.sheetnames == ["Resumo", "Itens", "Recomendacoes"]
    wb.close()


def test_export_xlsx_uses_selected_report_language():
    payload = export_xlsx(
        DummyParecer(idioma_relatorio="en"),
        [DummyItem(status="B")],
        [DummyRecomendacao()],
    )
    wb = load_workbook(io.BytesIO(payload), data_only=True)
    assert wb.sheetnames == ["Summary", "Items", "Recommendations"]
    assert wb["Summary"]["A1"].value == "ENGINEERING TECHNICAL OPINION"
    assert wb["Items"]["C2"].value == "APPROVED WITH COMMENTS"
    wb.close()


def test_export_docx_generates_document_xml():
    payload = export_docx(DummyParecer(), [DummyItem()], [DummyRecomendacao()])
    with zipfile.ZipFile(io.BytesIO(payload), "r") as archive:
        assert "word/document.xml" in archive.namelist()


def test_build_observacao_for_status_a_uses_default_when_empty():
    item = DummyItem(status="A", justificativa_tecnica="")
    observacao = _build_observacao(item)
    assert "aderente aos requisitos tecnicos" in observacao


@dataclass
class DummyCartaItem(DummyItem):
    id: uuid.UUID = field(default_factory=uuid.uuid4)


def test_carta_pendencias_includes_values_and_references():
    item = DummyCartaItem(
        numero=3,
        status="B",
        valor_requerido='workstation 19" rack, 4 monitores 55"',
        valor_fornecedor="Torre Dell T5860XL, 4 monitores 55\"",
        referencia_engenharia="MR-269=A, Cap. 2, Item 03",
        referencia_fornecedor="Proposta R3.2, Pag. 35",
    )
    wb = load_workbook(io.BytesIO(export_carta_pendencias(DummyParecer(), [item])))
    ws = wb["Pendencias"]

    # Cabeçalhos: layout enriquecido com valores e as duas referências.
    headers = [ws.cell(row=2, column=c).value for c in range(1, len(_CARTA_HEADERS) + 1)]
    assert headers == _CARTA_HEADERS
    assert "Prioridade" not in headers
    assert "Valor Requerido" in headers
    assert "Referencia (Engenharia)" in headers
    assert "Referencia (Fornecedor)" in headers

    # Linha de dados (3) carrega valores e referências reais.
    assert ws.cell(row=3, column=3).value == "AC"
    assert ws.cell(row=3, column=5).value == 'workstation 19" rack, 4 monitores 55"'
    assert ws.cell(row=3, column=6).value == "MR-269=A, Cap. 2, Item 03"
    assert ws.cell(row=3, column=8).value == "Proposta R3.2, Pag. 35"
    assert ws.cell(row=5, column=1).value == "Legenda de status"
    assert ws.cell(row=6, column=1).value == "AP"
    assert ws.cell(row=7, column=1).value == "AC"
    assert ws.cell(row=8, column=1).value == "RJ"
    assert ws.cell(row=9, column=1).value == "IA"
    wb.close()


def test_carta_pendencias_response_editable_and_item_id_roundtrip():
    # O reimport determinístico precisa que a coluna de resposta e o ITEM_ID fiquem
    # exatamente onde o exporter os colocou — guarda contra divergência futura.
    from app.api.v1.endpoints.ciclo_avaliativo import _COL_ITEM_ID, _COL_RESPOSTA

    assert _COL_RESPOSTA == _CARTA_COL_RESPOSTA
    assert _COL_ITEM_ID == _CARTA_COL_ITEM_ID

    item = DummyCartaItem(numero=1)
    wb = load_workbook(io.BytesIO(export_carta_pendencias(DummyParecer(), [item])))
    ws = wb["Pendencias"]

    assert ws.cell(row=2, column=_CARTA_COL_RESPOSTA).value == "Resposta do Fornecedor"
    assert ws.cell(row=2, column=_CARTA_COL_ITEM_ID).value == "ITEM_ID"
    # Só a coluna de resposta é editável; o requisito fica protegido.
    assert ws.cell(row=3, column=_CARTA_COL_RESPOSTA).protection.locked is False
    assert ws.cell(row=3, column=5).protection.locked is True
    # A protecao preserva o conteudo, mas permite ajustes visuais pelo fornecedor.
    assert ws.protection.sheet is True
    assert ws.protection.formatCells is False
    assert ws.protection.formatColumns is False
    assert ws.protection.formatRows is False
    assert ws.protection.autoFilter is False
    assert ws.protection.sort is False
    assert ws.protection.insertRows is True
    assert ws.protection.insertColumns is True
    assert ws.protection.deleteRows is True
    assert ws.protection.deleteColumns is True
    assert ws.auto_filter.ref == f"A2:{ws.cell(row=2, column=_CARTA_COL_ITEM_ID).column_letter}3"
    # ITEM_ID oculto e fiel ao item — o que torna o reimport determinístico.
    assert str(ws.cell(row=3, column=_CARTA_COL_ITEM_ID).value) == str(item.id)
    wb.close()


def test_reimport_columns_are_resolved_by_header_for_new_and_legacy_layout():
    from app.api.v1.endpoints.ciclo_avaliativo import _resolve_reimport_columns

    item = DummyCartaItem(numero=1)
    wb = load_workbook(io.BytesIO(export_carta_pendencias(DummyParecer(), [item])))
    ws = wb["Pendencias"]
    assert _resolve_reimport_columns(ws) == (_CARTA_COL_RESPOSTA, _CARTA_COL_ITEM_ID)
    wb.close()

    legacy_wb = Workbook()
    legacy_ws = legacy_wb.active
    legacy_ws.cell(row=2, column=11, value="Resposta do Fornecedor")
    legacy_ws.cell(row=2, column=12, value="ITEM_ID")
    assert _resolve_reimport_columns(legacy_ws) == (11, 12)

"""
Endpoints do ciclo com o fornecedor (fase CICLO_FORNECEDOR do caso técnico).

GET  /{parecer_id}/exportar/carta-pendencias
POST /{parecer_id}/ciclo/iniciar                    (W2: envia o caso ao ciclo)
POST /{parecer_id}/reimportar-respostas
GET  /{parecer_id}/ciclo/resumo
GET  /{parecer_id}/ciclo/reavaliacao
POST /{parecer_id}/ciclo/itens/{item_id}/decidir    (W4: decisão humana por item)
GET  /{parecer_id}/ciclo/itens/{item_id}/historico
"""
import asyncio
import io
import logging
import uuid
from datetime import datetime

from fastapi import APIRouter, Depends, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from pydantic import BaseModel
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.core.deps import get_current_user, require_role
from app.models.documento import Documento
from app.models.item_parecer import ItemParecer
from app.models.parecer import Parecer
from app.models.rodada_avaliacao import RodadaAvaliacao
from app.models.rodada_fornecedor import TIPOS_RODADA, RodadaFornecedor
from app.services.audit import registrar_auditoria
from app.services.evaluator import avaliar_resposta
from app.services.exporter import (
    _CARTA_COL_ITEM_ID,
    _CARTA_COL_RESPOSTA,
    export_carta_pendencias,
    export_ciclo_rodada,
)
from app.services.state_machine import (
    ACEITO,
    ANALISE,
    CICLO_FORNECEDOR,
    DECISOES_HUMANAS,
    DESATIVADO,
    EM_REAVALIACAO,
    PENDENTE_FORNECEDOR,
    VERIFICACAO_FINAL,
    FECHADO,
    compute_avanco_automatico,
    compute_resumo_ciclo,
    evento_para_decisao,
    todos_aceitos,
    transicionar,
)

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/pareceres", tags=["ciclo-avaliativo"])

# Posições das colunas na carta de pendências (1-based). Fonte única: o exporter —
# importadas para o reimport nunca divergir do layout gerado.
_COL_RESPOSTA = _CARTA_COL_RESPOSTA
_COL_ITEM_ID = _CARTA_COL_ITEM_ID
# Linha onde começam os dados (1=instrução, 2=cabeçalho, 3+=dados)
_DATA_START_ROW = 3


def _find_header_column(ws, header: str, fallback: int) -> int:
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=2, column=col).value
        if str(value or "").strip() == header:
            return col
    return fallback


def _resolve_reimport_columns(ws) -> tuple[int, int]:
    return (
        _find_header_column(ws, "Resposta do Fornecedor", _COL_RESPOSTA),
        _find_header_column(ws, "ITEM_ID", _COL_ITEM_ID),
    )


# ──────────────────────────────────────────────────────────────────────────────
# 2B — Export carta de pendências
# ──────────────────────────────────────────────────────────────────────────────

@router.get("/{parecer_id}/exportar/carta-pendencias")
async def exportar_carta_pendencias(
    parecer_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    parecer_result = await db.execute(select(Parecer).where(Parecer.id == parecer_id))
    parecer = parecer_result.scalar_one_or_none()
    if not parecer:
        raise HTTPException(status_code=404, detail="Parecer nao encontrado")

    itens_result = await db.execute(
        select(ItemParecer)
        .where(
            ItemParecer.parecer_id == parecer_id,
            ItemParecer.estado == PENDENTE_FORNECEDOR,
        )
        .order_by(ItemParecer.numero)
    )
    itens_pendentes = itens_result.scalars().all()

    if not itens_pendentes:
        raise HTTPException(
            status_code=400,
            detail="Nenhum item com estado PENDENTE_FORNECEDOR. Nada a exportar.",
        )

    # Rodada derivada do histórico (a coluna rodada_atual foi removida)
    rodada_result = await db.execute(
        select(func.max(RodadaAvaliacao.numero_rodada))
        .join(ItemParecer, ItemParecer.id == RodadaAvaliacao.item_id)
        .where(ItemParecer.parecer_id == parecer_id)
    )
    rodada = rodada_result.scalar() or 1

    try:
        content = export_carta_pendencias(parecer, itens_pendentes, rodada=rodada)
    except Exception:
        logger.exception("Falha ao exportar carta de pendencias para parecer %s", parecer_id)
        raise HTTPException(status_code=500, detail="Erro ao gerar carta de pendencias")
    filename = f"carta_pendencias_{parecer.numero_parecer}_R{rodada}.xlsx"
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ──────────────────────────────────────────────────────────────────────────────
# 2C — Reimport respostas do fornecedor
# ──────────────────────────────────────────────────────────────────────────────

class ItemProcessado(BaseModel):
    item_id: str
    numero: int
    veredito_ia: str
    justificativa_ia: str
    acao_requerida: str | None


class ItemIgnorado(BaseModel):
    linha: int
    motivo: str
    item_id_raw: str | None = None


class ReimportResult(BaseModel):
    processados: list[ItemProcessado]
    ignorados: list[ItemIgnorado]
    total_processados: int
    total_ignorados: int
    mensagem: str


@router.post("/{parecer_id}/reimportar-respostas", response_model=ReimportResult)
async def reimportar_respostas(
    parecer_id: uuid.UUID,
    arquivo: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    _current_user=Depends(require_role("admin", "analista")),
):
    if not (arquivo.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Apenas arquivos .xlsx sao aceitos.")

    parecer_result = await db.execute(select(Parecer).where(Parecer.id == parecer_id))
    parecer = parecer_result.scalar_one_or_none()
    if not parecer:
        raise HTTPException(status_code=404, detail="Parecer nao encontrado")
    if parecer.status_processamento == "processando":
        raise HTTPException(status_code=400, detail="Parecer em processamento. Aguarde.")
    if parecer.fase_caso != CICLO_FORNECEDOR:
        raise HTTPException(
            status_code=400,
            detail=(
                f"Respostas do fornecedor so podem ser carregadas na fase "
                f"CICLO_FORNECEDOR (fase atual: {parecer.fase_caso}). "
                "Inicie o ciclo apos gerar o parecer."
            ),
        )

    # Lê o Excel em memória
    raw_bytes = await arquivo.read()
    try:
        wb = load_workbook(io.BytesIO(raw_bytes), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Arquivo Excel invalido ou corrompido.")

    ws = wb.active
    col_resposta, col_item_id = _resolve_reimport_columns(ws)

    # Carrega todos os itens do parecer indexados por UUID para lookup O(1)
    itens_result = await db.execute(
        select(ItemParecer).where(ItemParecer.parecer_id == parecer_id)
    )
    itens_por_id: dict[str, ItemParecer] = {
        str(i.id): i for i in itens_result.scalars().all()
    }

    # Carrega a última rodada de cada item para obter a pendência (acao_requerida)
    rodadas_result = await db.execute(
        select(RodadaAvaliacao)
        .where(RodadaAvaliacao.item_id.in_([uuid.UUID(k) for k in itens_por_id]))
        .order_by(RodadaAvaliacao.numero_rodada)
    )
    ultima_rodada_por_item: dict[str, RodadaAvaliacao] = {}
    for rodada in rodadas_result.scalars().all():
        ultima_rodada_por_item[str(rodada.item_id)] = rodada

    now = datetime.utcnow()
    processados: list[ItemProcessado] = []
    ignorados: list[ItemIgnorado] = []

    # A carta XLSX é o caminho determinístico do Tipo 2: os vínculos vêm do
    # ITEM_ID da planilha, já confirmados — sem vinculação LLM.
    proxima_rodada_forn = await db.scalar(
        select(func.max(RodadaFornecedor.numero)).where(
            RodadaFornecedor.parecer_id == parecer_id
        )
    )
    rodada_fornecedor = RodadaFornecedor(
        parecer_id=parecer_id,
        numero=(proxima_rodada_forn or 0) + 1,
        tipo="RESPOSTA_ITENS",
        status="AVALIADA",
        criado_em=now,
    )
    db.add(rodada_fornecedor)
    await db.flush()

    for row_num in range(_DATA_START_ROW, ws.max_row + 1):
        resposta_cell = ws.cell(row=row_num, column=col_resposta)
        item_id_cell = ws.cell(row=row_num, column=col_item_id)

        resposta_raw = resposta_cell.value
        item_id_raw = str(item_id_cell.value).strip() if item_id_cell.value else None

        # Linha vazia — ignora silenciosamente
        if not item_id_raw and not resposta_raw:
            continue

        # ITEM_ID ausente ou inválido → resolução manual
        if not item_id_raw:
            ignorados.append(ItemIgnorado(
                linha=row_num,
                motivo="ITEM_ID ausente. Nao e possivel determinar o item correspondente.",
                item_id_raw=None,
            ))
            continue

        try:
            uuid.UUID(item_id_raw)
        except ValueError:
            ignorados.append(ItemIgnorado(
                linha=row_num,
                motivo=f"ITEM_ID invalido (nao e UUID): '{item_id_raw}'",
                item_id_raw=item_id_raw,
            ))
            continue

        if item_id_raw not in itens_por_id:
            ignorados.append(ItemIgnorado(
                linha=row_num,
                motivo=f"ITEM_ID nao pertence a este parecer: '{item_id_raw}'",
                item_id_raw=item_id_raw,
            ))
            continue

        item = itens_por_id[item_id_raw]

        # Resposta em branco → sem ação
        resposta = str(resposta_raw).strip() if resposta_raw else ""
        if not resposta:
            ignorados.append(ItemIgnorado(
                linha=row_num,
                motivo="Resposta em branco. Item nao processado.",
                item_id_raw=item_id_raw,
            ))
            continue

        # Item não está aguardando resposta → alerta, mas processa mesmo assim
        if item.estado != PENDENTE_FORNECEDOR:
            ignorados.append(ItemIgnorado(
                linha=row_num,
                motivo=(
                    f"Item {item.numero} esta em estado '{item.estado}', nao em "
                    f"PENDENTE_FORNECEDOR. Resposta ignorada para evitar inconsistencia."
                ),
                item_id_raw=item_id_raw,
            ))
            continue

        # Determina o número da nova rodada
        ultima_rodada = ultima_rodada_por_item.get(item_id_raw)
        nova_rodada_num = (ultima_rodada.numero_rodada + 1) if ultima_rodada else 2

        # Pendência da rodada anterior (fallback: campo direto do item)
        pendencia = ""
        if ultima_rodada and ultima_rodada.acao_requerida:
            pendencia = ultima_rodada.acao_requerida
        elif item.acao_requerida:
            pendencia = item.acao_requerida
        else:
            pendencia = item.descricao_requisito or ""

        # Chama o Agente Avaliador (síncrono → thread)
        evaluation = await asyncio.to_thread(
            avaliar_resposta,
            item.descricao_requisito or "",
            pendencia,
            resposta,
        )

        # Cria a nova RodadaAvaliacao (append-only)
        nova_rodada = RodadaAvaliacao(
            id=uuid.uuid4(),
            item_id=item.id,
            rodada_fornecedor_id=rodada_fornecedor.id,
            numero_rodada=nova_rodada_num,
            origem="RESPOSTA_FORNECEDOR",
            conteudo=resposta,
            trecho_vinculado=resposta,
            vinculo_confianca="ALTA",
            vinculo_metodo="DETERMINISTICO",
            veredito_ia=evaluation.veredito,
            justificativa_ia=evaluation.justificativa,
            acao_requerida=evaluation.acao_requerida,
            criado_em=now,
        )
        db.add(nova_rodada)

        # Transiciona o estado do item para EM_REAVALIACAO
        item.estado = transicionar(item.estado, "fornecedor_respondeu")

        processados.append(ItemProcessado(
            item_id=item_id_raw,
            numero=item.numero,
            veredito_ia=evaluation.veredito,
            justificativa_ia=evaluation.justificativa,
            acao_requerida=evaluation.acao_requerida,
        ))

        logger.info(
            "Reimport: parecer=%s item=%s rodada=%d veredito=%s",
            parecer_id, item_id_raw, nova_rodada_num, evaluation.veredito,
        )

    await db.commit()

    n_proc = len(processados)
    n_ign = len(ignorados)
    mensagem = (
        f"{n_proc} item(ns) processado(s) com sucesso. "
        f"{n_ign} ignorado(s) (ver campo 'ignorados' para detalhes)."
    )
    if n_proc == 0:
        mensagem = "Nenhum item foi processado. Verifique se o arquivo esta correto e as respostas preenchidas."

    return ReimportResult(
        processados=processados,
        ignorados=ignorados,
        total_processados=n_proc,
        total_ignorados=n_ign,
        mensagem=mensagem,
    )


# ──────────────────────────────────────────────────────────────────────────────
# Fase 3 — Validação humana e fechamento do ciclo
# ──────────────────────────────────────────────────────────────────────────────

class EstadoCount(BaseModel):
    estado: str
    total: int


class CicloResumoResponse(BaseModel):
    fase_caso: str
    desfecho: str | None
    contagem_por_estado: list[EstadoCount]
    total_itens: int
    tem_pendentes: bool
    tem_em_reavaliacao: bool
    todos_aceitos: bool


class RodadaResponse(BaseModel):
    id: str
    numero_rodada: int
    origem: str
    conteudo: str | None
    anexo_ref: str | None
    classificacao_ia: str | None
    veredito_ia: str | None
    justificativa_ia: str | None
    acao_requerida: str | None
    decisao_humana: str | None
    revisor: str | None
    criado_em: str


class ItemRevisaoResponse(BaseModel):
    id: str
    numero: int
    categoria: str | None
    descricao_requisito: str
    valor_requerido: str | None
    prioridade: str | None
    estado: str
    ultima_rodada: RodadaResponse | None


def _rodada_to_response(r: RodadaAvaliacao) -> RodadaResponse:
    return RodadaResponse(
        id=str(r.id),
        numero_rodada=r.numero_rodada,
        origem=r.origem,
        conteudo=r.conteudo,
        anexo_ref=r.anexo_ref,
        classificacao_ia=r.classificacao_ia,
        veredito_ia=r.veredito_ia,
        justificativa_ia=r.justificativa_ia,
        acao_requerida=r.acao_requerida,
        decisao_humana=r.decisao_humana,
        revisor=r.revisor,
        criado_em=r.criado_em.isoformat(),
    )


@router.get("/{parecer_id}/ciclo/resumo", response_model=CicloResumoResponse)
async def ciclo_resumo(
    parecer_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    parecer_result = await db.execute(select(Parecer).where(Parecer.id == parecer_id))
    parecer = parecer_result.scalar_one_or_none()
    if not parecer:
        raise HTTPException(status_code=404, detail="Parecer nao encontrado")

    estados_result = await db.execute(
        select(ItemParecer.estado).where(ItemParecer.parecer_id == parecer_id)
    )
    todos_estados = [row[0] for row in estados_result.all()]

    from collections import Counter
    counts = Counter(todos_estados)
    contagem = [
        EstadoCount(estado=estado, total=total)
        for estado, total in sorted(counts.items())
    ]
    resumo = compute_resumo_ciclo(todos_estados)

    return CicloResumoResponse(
        fase_caso=parecer.fase_caso,
        desfecho=parecer.desfecho,
        contagem_por_estado=contagem,
        total_itens=resumo["total_itens"],
        tem_pendentes=counts.get(PENDENTE_FORNECEDOR, 0) > 0,
        tem_em_reavaliacao=counts.get(EM_REAVALIACAO, 0) > 0,
        todos_aceitos=resumo["todos_aceitos"],
    )


@router.get("/{parecer_id}/ciclo/reavaliacao", response_model=list[ItemRevisaoResponse])
async def itens_em_reavaliacao(
    parecer_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    itens_result = await db.execute(
        select(ItemParecer)
        .where(
            ItemParecer.parecer_id == parecer_id,
            ItemParecer.estado == EM_REAVALIACAO,
        )
        .order_by(ItemParecer.numero)
    )
    itens = itens_result.scalars().all()

    if not itens:
        return []

    item_ids = [i.id for i in itens]
    rodadas_result = await db.execute(
        select(RodadaAvaliacao)
        .where(RodadaAvaliacao.item_id.in_(item_ids))
        .order_by(RodadaAvaliacao.numero_rodada)
    )
    ultimas: dict[str, RodadaAvaliacao] = {}
    for r in rodadas_result.scalars().all():
        ultimas[str(r.item_id)] = r

    return [
        ItemRevisaoResponse(
            id=str(i.id),
            numero=i.numero,
            categoria=i.categoria,
            descricao_requisito=i.descricao_requisito,
            valor_requerido=i.valor_requerido,
            prioridade=i.prioridade,
            estado=i.estado,
            ultima_rodada=_rodada_to_response(ultimas[str(i.id)]) if str(i.id) in ultimas else None,
        )
        for i in itens
    ]


@router.get("/{parecer_id}/ciclo/itens", response_model=list[ItemRevisaoResponse])
async def itens_do_ciclo(
    parecer_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """Todos os itens do caso com a ULTIMA rodada — visão de tabela completa do
    ciclo (decididos + pendentes), para o engenheiro ver tudo de uma vez. Difere
    de /ciclo/reavaliacao, que traz apenas os itens em EM_REAVALIACAO."""
    itens_result = await db.execute(
        select(ItemParecer)
        .where(
            ItemParecer.parecer_id == parecer_id,
            ItemParecer.estado != DESATIVADO,
        )
        .order_by(ItemParecer.numero)
    )
    itens = itens_result.scalars().all()
    if not itens:
        return []

    item_ids = [i.id for i in itens]
    rodadas_result = await db.execute(
        select(RodadaAvaliacao)
        .where(RodadaAvaliacao.item_id.in_(item_ids))
        .order_by(RodadaAvaliacao.numero_rodada)
    )
    ultimas: dict[str, RodadaAvaliacao] = {}
    for r in rodadas_result.scalars().all():
        ultimas[str(r.item_id)] = r  # ordenado: a maior rodada sobrescreve

    return [
        ItemRevisaoResponse(
            id=str(i.id),
            numero=i.numero,
            categoria=i.categoria,
            descricao_requisito=i.descricao_requisito,
            valor_requerido=i.valor_requerido,
            prioridade=i.prioridade,
            estado=i.estado,
            ultima_rodada=_rodada_to_response(ultimas[str(i.id)])
            if str(i.id) in ultimas
            else None,
        )
        for i in itens
    ]


@router.get("/{parecer_id}/exportar/ciclo-rodada")
async def exportar_ciclo_rodada(
    parecer_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """Excel com a visao completa da rodada (todos os itens + resposta + avaliacao
    IA + decisao/estado) — para o engenheiro revisar tudo de uma vez."""
    parecer_result = await db.execute(select(Parecer).where(Parecer.id == parecer_id))
    parecer = parecer_result.scalar_one_or_none()
    if not parecer:
        raise HTTPException(status_code=404, detail="Parecer nao encontrado")

    itens_result = await db.execute(
        select(ItemParecer)
        .where(
            ItemParecer.parecer_id == parecer_id,
            ItemParecer.estado != DESATIVADO,
        )
        .order_by(ItemParecer.numero)
    )
    itens = itens_result.scalars().all()
    if not itens:
        raise HTTPException(status_code=400, detail="Nenhum item no ciclo para exportar.")

    item_ids = [i.id for i in itens]
    rodadas_result = await db.execute(
        select(RodadaAvaliacao)
        .where(RodadaAvaliacao.item_id.in_(item_ids))
        .order_by(RodadaAvaliacao.numero_rodada)
    )
    ultimas: dict[str, RodadaAvaliacao] = {}
    for r in rodadas_result.scalars().all():
        ultimas[str(r.item_id)] = r
    rodada = max((r.numero_rodada for r in ultimas.values()), default=1)

    linhas = []
    for i in itens:
        ur = ultimas.get(str(i.id))
        linhas.append({
            "numero": i.numero,
            "categoria": i.categoria,
            "prioridade": i.prioridade,
            "descricao_requisito": i.descricao_requisito,
            "valor_requerido": i.valor_requerido,
            "resposta_fornecedor": ur.conteudo if ur else None,
            "veredito_ia": ur.veredito_ia if ur else None,
            "justificativa_ia": ur.justificativa_ia if ur else None,
            "decisao_humana": ur.decisao_humana if ur else None,
            "estado": i.estado,
        })

    try:
        content = export_ciclo_rodada(parecer, linhas, rodada=rodada)
    except Exception:
        logger.exception("Falha ao exportar visao da rodada para parecer %s", parecer_id)
        raise HTTPException(status_code=500, detail="Erro ao gerar Excel da rodada")
    filename = f"ciclo_rodada_{parecer.numero_parecer}_R{rodada}.xlsx"
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


class IniciarCicloResponse(BaseModel):
    fase_caso: str
    mensagem: str


@router.post("/{parecer_id}/ciclo/iniciar", response_model=IniciarCicloResponse)
async def iniciar_ciclo(
    parecer_id: uuid.UUID,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(require_role("admin", "analista")),
):
    """
    Operação W2 (bloco 17-18 do fluxo): o engenheiro validou a análise e gera o
    parecer para envio ao fornecedor — o caso avança de ANALISE para
    CICLO_FORNECEDOR. Se todos os itens já estão aceitos (100% A), pula direto
    para VERIFICACAO_FINAL: não há nada a tratar com o fornecedor.
    """
    parecer_result = await db.execute(select(Parecer).where(Parecer.id == parecer_id))
    parecer = parecer_result.scalar_one_or_none()
    if not parecer:
        raise HTTPException(status_code=404, detail="Parecer nao encontrado")
    if parecer.fase_caso != ANALISE:
        raise HTTPException(
            status_code=400,
            detail=f"Ciclo so pode ser iniciado na fase ANALISE (atual: {parecer.fase_caso}).",
        )
    if parecer.status_processamento != "concluido" or (parecer.total_itens or 0) == 0:
        raise HTTPException(
            status_code=400,
            detail="Execute a analise antes de iniciar o ciclo com o fornecedor.",
        )

    estados_result = await db.execute(
        select(ItemParecer.estado).where(ItemParecer.parecer_id == parecer_id)
    )
    estados = [row[0] for row in estados_result.all()]

    if todos_aceitos(estados):
        parecer.fase_caso = VERIFICACAO_FINAL
        mensagem = (
            "Todos os itens estao aceitos — caso avancou direto para a "
            "verificacao final."
        )
    else:
        parecer.fase_caso = CICLO_FORNECEDOR
        mensagem = (
            "Ciclo com o fornecedor iniciado. Exporte a carta de pendencias e "
            "carregue as respostas quando recebidas."
        )

    await registrar_auditoria(
        db, current_user, "w2_iniciar_ciclo", "parecer",
        recurso_id=str(parecer_id),
        detalhes=f"fase_caso={parecer.fase_caso}",
        request=request,
    )
    await db.commit()

    return IniciarCicloResponse(fase_caso=parecer.fase_caso, mensagem=mensagem)


class DecisoHumanaRequest(BaseModel):
    decisao: str  # ACEITAR | ESCLARECER | REJEITAR | REPROVAR_CASO
    comentario: str | None = None


class DecisoHumanaResponse(BaseModel):
    item_id: str
    numero: int
    novo_estado: str
    fase_caso: str
    desfecho: str | None
    mensagem: str


@router.post("/{parecer_id}/ciclo/itens/{item_id}/decidir", response_model=DecisoHumanaResponse)
async def decidir_item(
    parecer_id: uuid.UUID,
    item_id: uuid.UUID,
    payload: DecisoHumanaRequest,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(require_role("admin", "analista")),
):
    """
    Operação W4 (blocos 25-26 do fluxo): decisão humana por item.

      ACEITAR       → item conforme (ACEITO)
      ESCLARECER    → falta informação; volta ao fornecedor
      REJEITAR      → incorreto; fornecedor refaz
      REPROVAR_CASO → item crítico encerra o caso inteiro (desfecho REPROVADO)

    Quando todos os itens ativos ficam ACEITOS, o caso avança automaticamente
    para VERIFICACAO_FINAL (sem gate manual).
    """
    if payload.decisao not in DECISOES_HUMANAS:
        raise HTTPException(
            status_code=400,
            detail=f"decisao deve ser uma de: {', '.join(sorted(DECISOES_HUMANAS))}.",
        )

    parecer_result = await db.execute(select(Parecer).where(Parecer.id == parecer_id))
    parecer = parecer_result.scalar_one_or_none()
    if not parecer:
        raise HTTPException(status_code=404, detail="Parecer nao encontrado")
    if parecer.fase_caso != CICLO_FORNECEDOR:
        raise HTTPException(
            status_code=400,
            detail=f"Decisoes por item so na fase CICLO_FORNECEDOR (atual: {parecer.fase_caso}).",
        )

    item_result = await db.execute(
        select(ItemParecer).where(
            ItemParecer.id == item_id,
            ItemParecer.parecer_id == parecer_id,
        )
    )
    item = item_result.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="Item nao encontrado neste parecer.")
    if item.estado != EM_REAVALIACAO:
        raise HTTPException(
            status_code=400,
            detail=f"Item esta em estado '{item.estado}', nao em EM_REAVALIACAO.",
        )
    status_anterior = item.status
    estado_anterior = item.estado

    # Registra a decisão (W4) na rodada mais recente
    ultima_result = await db.execute(
        select(RodadaAvaliacao)
        .where(RodadaAvaliacao.item_id == item_id)
        .order_by(RodadaAvaliacao.numero_rodada.desc())
        .limit(1)
    )
    ultima_rodada = ultima_result.scalar_one_or_none()
    if ultima_rodada:
        ultima_rodada.decisao_humana = payload.decisao
        revisor_id = getattr(current_user, "id", None)
        ultima_rodada.revisor = str(revisor_id) if revisor_id else None
        if payload.comentario:
            ultima_rodada.conteudo = (
                f"{ultima_rodada.conteudo}\n\n[Comentario da decisao] {payload.comentario}"
                if ultima_rodada.conteudo
                else f"[Comentario da decisao] {payload.comentario}"
            )

    item.estado = transicionar(item.estado, evento_para_decisao(payload.decisao))
    # Highlight da revisao de spec se encerra quando o item e reavaliado
    item.marcacao_revisao = None

    # REPROVAR_CASO: um item crítico encerra o caso inteiro (bloco 25 → terminal)
    if payload.decisao == "REPROVAR_CASO":
        parecer.fase_caso = FECHADO
        parecer.desfecho = "REPROVADO"
        parecer.fechado_em = datetime.utcnow()
        parecer.fechado_por = getattr(current_user, "id", None)
        parecer.motivo_fechamento = (
            payload.comentario
            or f"Caso reprovado pelo item critico {item.numero}: {item.descricao_requisito[:200]}"
        )
        mensagem = f"Caso REPROVADO pelo item critico {item.numero}."
    else:
        # Avanço automático: todos os itens ativos aceitos → VERIFICACAO_FINAL
        estados_result = await db.execute(
            select(ItemParecer.estado).where(ItemParecer.parecer_id == parecer_id)
        )
        estados = [row[0] for row in estados_result.all()]
        nova_fase = compute_avanco_automatico(parecer.fase_caso, estados)
        if nova_fase:
            parecer.fase_caso = nova_fase
            mensagem = (
                f"Item {item.numero} aceito. Todos os itens estao aceitos — "
                "caso avancou automaticamente para a verificacao final."
            )
        elif item.estado == ACEITO:
            mensagem = f"Item {item.numero} aceito."
        else:
            mensagem = (
                f"Item {item.numero} voltou ao fornecedor "
                f"({'esclarecimento' if payload.decisao == 'ESCLARECER' else 'refazer'})."
            )

    await registrar_auditoria(
        db, current_user, "w4_decidir_item", "item",
        recurso_id=str(item_id),
        detalhes=(
            f"item_numero={item.numero}; decisao={payload.decisao}; "
            f"status_anterior={status_anterior}; status_novo={item.status}; "
            f"estado_anterior={estado_anterior}; estado_novo={item.estado}"
        ),
        request=request,
    )
    await db.commit()

    return DecisoHumanaResponse(
        item_id=str(item_id),
        numero=item.numero,
        novo_estado=item.estado,
        fase_caso=parecer.fase_caso,
        desfecho=parecer.desfecho,
        mensagem=mensagem,
    )


@router.post(
    "/{parecer_id}/ciclo/itens/{item_id}/desfazer-decisao",
    response_model=DecisoHumanaResponse,
)
async def desfazer_decisao_item(
    parecer_id: uuid.UUID,
    item_id: uuid.UUID,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(require_role("admin", "analista")),
):
    """Desfaz a decisão W4 de um item (clique errado): volta o item para
    EM_REAVALIACAO e limpa a decisão da rodada. Cobre ACEITAR/ESCLARECER/REJEITAR;
    REPROVAR_CASO (que fecha o caso) NÃO é revertido aqui. Se o aceite havia
    avançado o caso para a verificação final, a fase regride para o ciclo."""
    parecer_result = await db.execute(select(Parecer).where(Parecer.id == parecer_id))
    parecer = parecer_result.scalar_one_or_none()
    if not parecer:
        raise HTTPException(status_code=404, detail="Parecer nao encontrado")
    if parecer.fase_caso not in (CICLO_FORNECEDOR, VERIFICACAO_FINAL):
        raise HTTPException(
            status_code=400,
            detail=f"So e possivel desfazer decisoes no ciclo (fase atual: {parecer.fase_caso}).",
        )

    item_result = await db.execute(
        select(ItemParecer).where(
            ItemParecer.id == item_id,
            ItemParecer.parecer_id == parecer_id,
        )
    )
    item = item_result.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="Item nao encontrado neste parecer.")

    ultima_result = await db.execute(
        select(RodadaAvaliacao)
        .where(RodadaAvaliacao.item_id == item_id)
        .order_by(RodadaAvaliacao.numero_rodada.desc())
        .limit(1)
    )
    ultima_rodada = ultima_result.scalar_one_or_none()
    if not ultima_rodada or not ultima_rodada.decisao_humana:
        raise HTTPException(status_code=400, detail="Nao ha decisao para desfazer neste item.")
    if ultima_rodada.decisao_humana == "REPROVAR_CASO":
        raise HTTPException(
            status_code=400,
            detail="A reprovacao do caso nao pode ser desfeita por aqui.",
        )
    if item.estado not in (ACEITO, PENDENTE_FORNECEDOR):
        raise HTTPException(
            status_code=400,
            detail=f"Item em '{item.estado}' nao pode ser revertido.",
        )

    estado_anterior = item.estado
    decisao_anterior = ultima_rodada.decisao_humana
    fase_anterior = parecer.fase_caso

    item.estado = transicionar(item.estado, "desfazer_decisao")
    ultima_rodada.decisao_humana = None
    ultima_rodada.revisor = None

    # Se o aceite tinha avancado o caso para a verificacao final, regride: agora
    # ha item pendente de novo.
    if parecer.fase_caso == VERIFICACAO_FINAL:
        parecer.fase_caso = CICLO_FORNECEDOR

    await registrar_auditoria(
        db, current_user, "w4_desfazer_decisao", "item",
        recurso_id=str(item_id),
        detalhes=(
            f"item_numero={item.numero}; decisao_desfeita={decisao_anterior}; "
            f"estado_anterior={estado_anterior}; estado_novo={item.estado}; "
            f"fase_anterior={fase_anterior}; fase_nova={parecer.fase_caso}"
        ),
        request=request,
    )
    await db.commit()

    return DecisoHumanaResponse(
        item_id=str(item_id),
        numero=item.numero,
        novo_estado=item.estado,
        fase_caso=parecer.fase_caso,
        desfecho=parecer.desfecho,
        mensagem=f"Decisao do item {item.numero} desfeita — voltou para a fila.",
    )


class AplicarAvaliacaoResponse(BaseModel):
    aceitos: int
    pendencias: int
    fase_caso: str
    desfecho: str | None
    mensagem: str


# Veredito da IA (R2) → decisão humana aplicada em lote ("Aplicar e seguir").
# ATENDE vira aceite; PARCIAL/NAO_ATENDE (e ausência de veredito) voltam ao
# fornecedor como pendência. REPROVAR_CASO nunca é automático.
_VEREDITO_PARA_DECISAO = {
    "ATENDE": "ACEITAR",
    "PARCIAL": "ESCLARECER",
    "NAO_ATENDE": "REJEITAR",
}


@router.post(
    "/{parecer_id}/ciclo/aplicar-avaliacao", response_model=AplicarAvaliacaoResponse
)
async def aplicar_avaliacao(
    parecer_id: uuid.UUID,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(require_role("admin", "analista")),
):
    """Filosofia 'resolvedora': em vez de o engenheiro aceitar item a item, a
    avaliação da JULIA (veredito R2) é aplicada em LOTE com um clique. Cada item
    em EM_REAVALIACAO recebe a decisão derivada do seu veredito (ATENDE→aceito;
    PARCIAL/NAO_ATENDE→volta ao fornecedor) e o caso avança automaticamente se
    todos ficarem aceitos. Correções pontuais ficam por conta do chat/desfazer."""
    parecer_result = await db.execute(select(Parecer).where(Parecer.id == parecer_id))
    parecer = parecer_result.scalar_one_or_none()
    if not parecer:
        raise HTTPException(status_code=404, detail="Parecer nao encontrado")
    if parecer.fase_caso != CICLO_FORNECEDOR:
        raise HTTPException(
            status_code=400,
            detail=f"Aplicar avaliacao so na fase CICLO_FORNECEDOR (atual: {parecer.fase_caso}).",
        )

    itens_result = await db.execute(
        select(ItemParecer)
        .where(
            ItemParecer.parecer_id == parecer_id,
            ItemParecer.estado == EM_REAVALIACAO,
        )
        .order_by(ItemParecer.numero)
    )
    itens = itens_result.scalars().all()
    if not itens:
        raise HTTPException(status_code=400, detail="Nenhum item pendente de decisao.")

    item_ids = [i.id for i in itens]
    rodadas_result = await db.execute(
        select(RodadaAvaliacao)
        .where(RodadaAvaliacao.item_id.in_(item_ids))
        .order_by(RodadaAvaliacao.numero_rodada)
    )
    ultimas: dict[str, RodadaAvaliacao] = {}
    for r in rodadas_result.scalars().all():
        ultimas[str(r.item_id)] = r

    revisor_id = getattr(current_user, "id", None)
    aceitos = 0
    pendencias = 0
    for item in itens:
        ur = ultimas.get(str(item.id))
        veredito = ur.veredito_ia if ur else None
        decisao = _VEREDITO_PARA_DECISAO.get(veredito, "ESCLARECER")
        if ur:
            ur.decisao_humana = decisao
            ur.revisor = str(revisor_id) if revisor_id else None
        item.estado = transicionar(item.estado, evento_para_decisao(decisao))
        item.marcacao_revisao = None
        if decisao == "ACEITAR":
            aceitos += 1
        else:
            pendencias += 1

    # Avanço automático: todos os itens ativos aceitos → VERIFICACAO_FINAL
    estados_result = await db.execute(
        select(ItemParecer.estado).where(ItemParecer.parecer_id == parecer_id)
    )
    estados = [row[0] for row in estados_result.all()]
    nova_fase = compute_avanco_automatico(parecer.fase_caso, estados)
    if nova_fase:
        parecer.fase_caso = nova_fase
        mensagem = (
            f"{aceitos} item(ns) aceito(s) pela avaliacao da JULIA. Todos atendem — "
            "caso avancou para a verificacao final."
        )
    else:
        mensagem = (
            f"Avaliacao aplicada: {aceitos} aceito(s), {pendencias} voltaram ao "
            "fornecedor como pendencia. Exporte a carta e envie quando quiser."
        )

    await registrar_auditoria(
        db, current_user, "w4_aplicar_avaliacao", "parecer",
        recurso_id=str(parecer_id),
        detalhes=f"aceitos={aceitos}; pendencias={pendencias}; fase={parecer.fase_caso}",
        request=request,
    )
    await db.commit()

    return AplicarAvaliacaoResponse(
        aceitos=aceitos,
        pendencias=pendencias,
        fase_caso=parecer.fase_caso,
        desfecho=parecer.desfecho,
        mensagem=mensagem,
    )


@router.get("/{parecer_id}/ciclo/itens/{item_id}/historico", response_model=list[RodadaResponse])
async def historico_item(
    parecer_id: uuid.UUID,
    item_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    item_result = await db.execute(
        select(ItemParecer).where(
            ItemParecer.id == item_id,
            ItemParecer.parecer_id == parecer_id,
        )
    )
    if not item_result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="Item nao encontrado neste parecer.")

    rodadas_result = await db.execute(
        select(RodadaAvaliacao)
        .where(RodadaAvaliacao.item_id == item_id)
        .order_by(RodadaAvaliacao.numero_rodada)
    )
    return [_rodada_to_response(r) for r in rodadas_result.scalars().all()]


# ──────────────────────────────────────────────────────────────────────────────
# Rodadas do fornecedor — tipos 1-4 (blocos 21-24 do fluxo) + vinculação W3
# ──────────────────────────────────────────────────────────────────────────────

class RodadaFornecedorResponse(BaseModel):
    id: str
    numero: int
    tipo: str
    status: str
    proposta_final: bool
    tem_texto_colado: bool
    documento_nome: str | None
    erro_detalhe: str | None
    criado_em: str


class VinculoResponse(BaseModel):
    id: str
    item_id: str
    item_numero: int
    item_descricao: str
    item_estado: str
    trecho: str | None
    confianca: str | None
    metodo: str | None
    veredito_ia: str | None
    justificativa_ia: str | None


class RodadaDetalheResponse(RodadaFornecedorResponse):
    vinculos: list[VinculoResponse]


class CriarRodadaResponse(BaseModel):
    rodada_id: str
    task_id: str
    mensagem: str


def _rodada_forn_to_response(r: RodadaFornecedor, documento_nome: str | None) -> dict:
    return {
        "id": str(r.id),
        "numero": r.numero,
        "tipo": r.tipo,
        "status": r.status,
        "proposta_final": r.proposta_final,
        "tem_texto_colado": bool(r.texto_colado and r.texto_colado.strip()),
        "documento_nome": documento_nome,
        "erro_detalhe": r.erro_detalhe,
        "criado_em": r.criado_em.isoformat(),
    }


async def _get_rodada(
    parecer_id: uuid.UUID, rodada_id: uuid.UUID, db: AsyncSession
) -> RodadaFornecedor:
    result = await db.execute(
        select(RodadaFornecedor).where(
            RodadaFornecedor.id == rodada_id,
            RodadaFornecedor.parecer_id == parecer_id,
        )
    )
    rodada = result.scalar_one_or_none()
    if not rodada:
        raise HTTPException(status_code=404, detail="Rodada nao encontrada neste parecer.")
    return rodada


@router.post(
    "/{parecer_id}/rodadas",
    response_model=CriarRodadaResponse,
    status_code=202,
)
async def criar_rodada(
    parecer_id: uuid.UUID,
    request: Request,
    tipo: str = Form(...),
    texto_colado: str | None = Form(None),
    proposta_final: bool = Form(False),
    arquivo: UploadFile | None = File(None),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(require_role("admin", "analista")),
):
    """
    Bloco 22 do fluxo: o engenheiro carrega a resposta do fornecedor informando
    o tipo (1=PROPOSTA_REVISADA, 2=RESPOSTA_ITENS,
    3=RESPOSTA_ITENS_PROPOSTA_POSTERIOR, 4=EMAIL_AVULSO). O material entra por
    upload (pdf/docx/xlsx) ou texto colado. A vinculação LLM roda em background.
    """
    if tipo not in TIPOS_RODADA:
        raise HTTPException(
            status_code=400,
            detail=f"tipo deve ser um de: {', '.join(sorted(TIPOS_RODADA))}.",
        )

    parecer_result = await db.execute(select(Parecer).where(Parecer.id == parecer_id))
    parecer = parecer_result.scalar_one_or_none()
    if not parecer:
        raise HTTPException(status_code=404, detail="Parecer nao encontrado")
    # Rodadas normais entram no CICLO_FORNECEDOR; a proposta revisada FINAL
    # (blocos 30-31) entra na VERIFICACAO_FINAL e não passa por vinculação.
    if proposta_final:
        if parecer.fase_caso != VERIFICACAO_FINAL:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"Proposta final so pode ser carregada na fase "
                    f"VERIFICACAO_FINAL (fase atual: {parecer.fase_caso})."
                ),
            )
    elif parecer.fase_caso != CICLO_FORNECEDOR:
        raise HTTPException(
            status_code=400,
            detail=(
                f"Rodadas so podem ser criadas na fase CICLO_FORNECEDOR "
                f"(fase atual: {parecer.fase_caso})."
            ),
        )

    tem_texto = bool(texto_colado and texto_colado.strip())
    tem_arquivo = arquivo is not None and bool(arquivo.filename)
    if not tem_texto and not tem_arquivo:
        raise HTTPException(
            status_code=400,
            detail="Envie um arquivo ou cole o texto da resposta do fornecedor.",
        )
    if tipo == "PROPOSTA_REVISADA" and not tem_arquivo:
        raise HTTPException(
            status_code=400,
            detail="Proposta revisada (Tipo 1) exige upload do documento da proposta.",
        )

    documento_id = None
    documento_nome = None
    if tem_arquivo:
        from app.api.v1.endpoints.documentos import _upload_doc

        doc_response = await _upload_doc(parecer_id, "resposta_fornecedor", arquivo, db)
        documento_id = uuid.UUID(doc_response.id)
        documento_nome = doc_response.nome_arquivo

    proxima = await db.scalar(
        select(func.max(RodadaFornecedor.numero)).where(
            RodadaFornecedor.parecer_id == parecer_id
        )
    )
    rodada = RodadaFornecedor(
        parecer_id=parecer_id,
        numero=(proxima or 0) + 1,
        tipo=tipo,
        texto_colado=texto_colado if tem_texto else None,
        documento_id=documento_id,
        proposta_final=proposta_final,
        status="RECEBIDA",
        criado_por=getattr(current_user, "id", None),
        criado_em=datetime.utcnow(),
    )
    db.add(rodada)
    await db.commit()
    await db.refresh(rodada)

    if proposta_final:
        # Proposta final não passa por vinculação — será usada pela verificação R3
        task_id = ""
        mensagem = (
            "Proposta final registrada. Execute a verificacao final para "
            "compara-la com os acordos do caso."
        )
    else:
        from app.services.ciclo import start_vinculacao_in_background

        task_id = start_vinculacao_in_background(str(rodada.id))
        mensagem = (
            "Rodada registrada. A LLM esta vinculando as respostas aos itens "
            "abertos — acompanhe pelo endpoint de progresso e confirme a "
            "vinculacao quando concluida."
        )

    await registrar_auditoria(
        db, current_user, "criar_rodada_fornecedor", "rodada_fornecedor",
        recurso_id=str(rodada.id),
        detalhes=f"tipo={tipo}, proposta_final={proposta_final}, arquivo={documento_nome or '-'}",
        request=request,
    )
    await db.commit()

    return CriarRodadaResponse(
        rodada_id=str(rodada.id),
        task_id=task_id,
        mensagem=mensagem,
    )


@router.get("/{parecer_id}/rodadas", response_model=list[RodadaFornecedorResponse])
async def listar_rodadas(
    parecer_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    result = await db.execute(
        select(RodadaFornecedor, Documento.nome_arquivo)
        .outerjoin(Documento, Documento.id == RodadaFornecedor.documento_id)
        .where(RodadaFornecedor.parecer_id == parecer_id)
        .order_by(RodadaFornecedor.numero)
    )
    return [
        RodadaFornecedorResponse(**_rodada_forn_to_response(r, nome))
        for r, nome in result.all()
    ]


@router.get("/{parecer_id}/rodadas/{rodada_id}", response_model=RodadaDetalheResponse)
async def detalhar_rodada(
    parecer_id: uuid.UUID,
    rodada_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    rodada = await _get_rodada(parecer_id, rodada_id, db)

    documento_nome = None
    if rodada.documento_id:
        documento_nome = await db.scalar(
            select(Documento.nome_arquivo).where(Documento.id == rodada.documento_id)
        )

    vinculos_result = await db.execute(
        select(RodadaAvaliacao, ItemParecer)
        .join(ItemParecer, ItemParecer.id == RodadaAvaliacao.item_id)
        .where(RodadaAvaliacao.rodada_fornecedor_id == rodada.id)
        .order_by(ItemParecer.numero)
    )
    vinculos = [
        VinculoResponse(
            id=str(av.id),
            item_id=str(item.id),
            item_numero=item.numero,
            item_descricao=item.descricao_requisito,
            item_estado=item.estado,
            trecho=av.trecho_vinculado,
            confianca=av.vinculo_confianca,
            metodo=av.vinculo_metodo,
            veredito_ia=av.veredito_ia,
            justificativa_ia=av.justificativa_ia,
        )
        for av, item in vinculos_result.all()
    ]

    return RodadaDetalheResponse(
        **_rodada_forn_to_response(rodada, documento_nome),
        vinculos=vinculos,
    )


@router.get("/{parecer_id}/rodadas/{rodada_id}/progresso")
async def progresso_rodada(
    parecer_id: uuid.UUID,
    rodada_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    from app.core.progress import get_progress

    rodada = await _get_rodada(parecer_id, rodada_id, db)
    progress = get_progress(f"rodada:{rodada_id}") or {}
    return {
        "status": rodada.status,
        "percent": progress.get("percent"),
        "message": progress.get("message"),
        "stage": progress.get("stage"),
    }


class CorrigirVinculoRequest(BaseModel):
    item_numero: int | None = None  # novo item; None mantém
    remover: bool = False


@router.patch(
    "/{parecer_id}/rodadas/{rodada_id}/vinculos/{avaliacao_id}",
    response_model=VinculoResponse | None,
)
async def corrigir_vinculo(
    parecer_id: uuid.UUID,
    rodada_id: uuid.UUID,
    avaliacao_id: uuid.UUID,
    payload: CorrigirVinculoRequest,
    db: AsyncSession = Depends(get_db),
    _current_user=Depends(require_role("admin", "analista")),
):
    """Correção manual de um vínculo sugerido — antes da confirmação (W3)."""
    rodada = await _get_rodada(parecer_id, rodada_id, db)
    if rodada.status != "VINCULACAO_SUGERIDA":
        raise HTTPException(
            status_code=400,
            detail=f"Vinculos so podem ser corrigidos com status VINCULACAO_SUGERIDA (atual: {rodada.status}).",
        )

    av_result = await db.execute(
        select(RodadaAvaliacao).where(
            RodadaAvaliacao.id == avaliacao_id,
            RodadaAvaliacao.rodada_fornecedor_id == rodada.id,
        )
    )
    avaliacao = av_result.scalar_one_or_none()
    if not avaliacao:
        raise HTTPException(status_code=404, detail="Vinculo nao encontrado nesta rodada.")

    if payload.remover:
        await db.delete(avaliacao)
        await db.commit()
        return None

    if payload.item_numero is None:
        raise HTTPException(status_code=422, detail="Informe item_numero ou remover=true.")

    novo_item_result = await db.execute(
        select(ItemParecer).where(
            ItemParecer.parecer_id == parecer_id,
            ItemParecer.numero == payload.item_numero,
        )
    )
    novo_item = novo_item_result.scalar_one_or_none()
    if not novo_item:
        raise HTTPException(status_code=404, detail=f"Item {payload.item_numero} nao encontrado.")
    if novo_item.estado != PENDENTE_FORNECEDOR:
        raise HTTPException(
            status_code=400,
            detail=f"Item {novo_item.numero} nao esta aguardando resposta (estado: {novo_item.estado}).",
        )

    # Reaponta o vínculo e recalcula o número da rodada para o novo item
    proxima = await db.scalar(
        select(func.max(RodadaAvaliacao.numero_rodada)).where(
            RodadaAvaliacao.item_id == novo_item.id
        )
    )
    avaliacao.item_id = novo_item.id
    avaliacao.numero_rodada = (proxima or 0) + 1
    avaliacao.vinculo_metodo = "MANUAL"
    avaliacao.vinculo_confianca = "ALTA"
    await db.commit()

    return VinculoResponse(
        id=str(avaliacao.id),
        item_id=str(novo_item.id),
        item_numero=novo_item.numero,
        item_descricao=novo_item.descricao_requisito,
        item_estado=novo_item.estado,
        trecho=avaliacao.trecho_vinculado,
        confianca=avaliacao.vinculo_confianca,
        metodo=avaliacao.vinculo_metodo,
        veredito_ia=avaliacao.veredito_ia,
        justificativa_ia=avaliacao.justificativa_ia,
    )


class ConfirmarVinculacaoResponse(BaseModel):
    rodada_id: str
    status: str
    itens_transicionados: int
    task_id: str
    mensagem: str


@router.post(
    "/{parecer_id}/rodadas/{rodada_id}/confirmar-vinculacao",
    response_model=ConfirmarVinculacaoResponse,
    status_code=202,
)
async def confirmar_vinculacao(
    parecer_id: uuid.UUID,
    rodada_id: uuid.UUID,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(require_role("admin", "analista")),
):
    """
    Operação W3: o engenheiro valida a vinculação sugerida. Os itens vinculados
    transicionam para EM_REAVALIACAO e a avaliação LLM (R2) roda em background.
    """
    rodada = await _get_rodada(parecer_id, rodada_id, db)
    if rodada.status != "VINCULACAO_SUGERIDA":
        raise HTTPException(
            status_code=400,
            detail=f"Rodada nao aguarda confirmacao (status: {rodada.status}).",
        )

    vinculos_result = await db.execute(
        select(RodadaAvaliacao).where(RodadaAvaliacao.rodada_fornecedor_id == rodada.id)
    )
    vinculos = vinculos_result.scalars().all()
    if not vinculos:
        raise HTTPException(
            status_code=400,
            detail="Nenhum vinculo nesta rodada. Corrija os vinculos ou descarte a rodada.",
        )

    transicionados = 0
    for avaliacao in vinculos:
        item_result = await db.execute(
            select(ItemParecer).where(ItemParecer.id == avaliacao.item_id)
        )
        item = item_result.scalar_one()
        if item.estado == PENDENTE_FORNECEDOR:
            item.estado = transicionar(item.estado, "fornecedor_respondeu")
            transicionados += 1

    rodada.status = "VINCULACAO_CONFIRMADA"

    await registrar_auditoria(
        db, current_user, "w3_confirmar_vinculacao", "rodada_fornecedor",
        recurso_id=str(rodada.id),
        detalhes=f"{len(vinculos)} vinculo(s), {transicionados} item(ns) em reavaliacao",
        request=request,
    )
    await db.commit()

    from app.services.ciclo import start_avaliacao_in_background

    task_id = start_avaliacao_in_background(str(rodada.id))

    return ConfirmarVinculacaoResponse(
        rodada_id=str(rodada.id),
        status=rodada.status,
        itens_transicionados=transicionados,
        task_id=task_id,
        mensagem=(
            "Vinculacao confirmada (W3). A LLM esta avaliando as respostas — "
            "os itens aparecerao para sua decisao quando concluir."
        ),
    )
